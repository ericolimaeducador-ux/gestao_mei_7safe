"""
Camada de dados — toda leitura/escrita acontece na planilha Excel.
A planilha tem 4 abas:
  - Lancamentos : receitas e despesas (livro-caixa)
  - DAS         : controle mensal do imposto
  - Notas       : notas fiscais emitidas
  - Cadastro    : dados do MEI (espelho do config, editável)

Usamos openpyxl (lê/escreve .xlsx sem precisar do Excel instalado).
A planilha é criada automaticamente na primeira execução.
"""

from datetime import datetime
from openpyxl import Workbook, load_workbook

from . import config

ABAS = {
    "Lancamentos": ["id", "data", "tipo", "descricao", "categoria", "valor", "nota_id"],
    "DAS": ["competencia", "valor", "vencimento", "status", "data_pagamento", "obs"],
    "Notas": ["id", "data", "tomador", "cnpj_cpf", "descricao", "valor", "tipo", "status", "link"],
    "Cadastro": ["chave", "valor"],
}


# ---------------------------------------------------------------------------
# Inicialização
# ---------------------------------------------------------------------------
def garantir_planilha():
    """Cria a planilha com as abas e cabeçalhos se ela ainda não existir."""
    if config.PLANILHA.exists():
        return
    wb = Workbook()
    wb.remove(wb.active)  # remove a aba padrão vazia
    for nome, cabecalho in ABAS.items():
        ws = wb.create_sheet(nome)
        ws.append(cabecalho)
    # preenche cadastro inicial a partir do config
    cad = wb["Cadastro"]
    for chave, valor in config.EMPRESA.items():
        cad.append([chave, str(valor)])
    wb.save(config.PLANILHA)


def _abrir():
    garantir_planilha()
    return load_workbook(config.PLANILHA)


def _proximo_id(ws):
    """Gera um id sequencial olhando a primeira coluna (ignora o cabeçalho)."""
    ids = [c.value for c in ws["A"][1:] if isinstance(c.value, int)]
    return (max(ids) + 1) if ids else 1


# ---------------------------------------------------------------------------
# Lançamentos (livro-caixa)
# ---------------------------------------------------------------------------
def listar_lancamentos():
    wb = _abrir()
    ws = wb["Lancamentos"]
    cols = ABAS["Lancamentos"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        out.append(dict(zip(cols, row)))
    return out


def adicionar_lancamento(data, tipo, descricao, categoria, valor, nota_id=""):
    """tipo = 'receita' ou 'despesa'."""
    wb = _abrir()
    ws = wb["Lancamentos"]
    novo_id = _proximo_id(ws)
    ws.append([novo_id, data, tipo, descricao, categoria, float(valor), nota_id])
    wb.save(config.PLANILHA)
    return novo_id


def remover_lancamento(lanc_id):
    wb = _abrir()
    ws = wb["Lancamentos"]
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == lanc_id:
            ws.delete_rows(i, 1)
            break
    wb.save(config.PLANILHA)


# ---------------------------------------------------------------------------
# DAS
# ---------------------------------------------------------------------------
def listar_das():
    wb = _abrir()
    ws = wb["DAS"]
    cols = ABAS["DAS"]
    return [dict(zip(cols, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]


def registrar_das(competencia, valor, vencimento, status="aberto", data_pagamento="", obs=""):
    """competencia = 'AAAA-MM'. Atualiza se já existir."""
    wb = _abrir()
    ws = wb["DAS"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == competencia:
            row[1].value, row[2].value, row[3].value = valor, vencimento, status
            row[4].value, row[5].value = data_pagamento, obs
            wb.save(config.PLANILHA)
            return
    ws.append([competencia, valor, vencimento, status, data_pagamento, obs])
    wb.save(config.PLANILHA)


# ---------------------------------------------------------------------------
# Notas fiscais
# ---------------------------------------------------------------------------
def listar_notas():
    wb = _abrir()
    ws = wb["Notas"]
    cols = ABAS["Notas"]
    return [dict(zip(cols, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]


def adicionar_nota(data, tomador, cnpj_cpf, descricao, valor, tipo, status, link=""):
    wb = _abrir()
    ws = wb["Notas"]
    novo_id = _proximo_id(ws)
    ws.append([novo_id, data, tomador, cnpj_cpf, descricao, float(valor), tipo, status, link])
    wb.save(config.PLANILHA)
    return novo_id
